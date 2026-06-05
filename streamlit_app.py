import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO
import re
import requests
from datetime import datetime
from PIL import Image, ImageDraw
import os
import json
try:
    from streamlit_image_coordinates import streamlit_image_coordinates
except ImportError:
    streamlit_image_coordinates = None


st.title("Quick ADOL Template Guide")

st.write("Insert numbers into the pre-formatted ADOL Scoresheet on Excel and Print it!")

# Step 1: Get Mentee's Name
st.subheader("1. Enter Mentee's name and today's date")
col1, col2 = st.columns(2)

with col1:
    mentee_name = st.text_input(
        "Enter name:",
        placeholder="John Doe"
    )

with col2:
    date_input = st.text_input(
        "Enter Today's date:",
        placeholder="MM/DD/YYYY"
    )

# Combine name and date
combined_name = f"{mentee_name} {date_input}".strip()

# Step 2: Choose input method
st.subheader("2. Choose Your Input Method")
input_method = st.radio(
    "How would you like to enter your scores?",
    options=["Type Numbers", "Click on Image"],
    key="input_method"
)

# Coordinate mappings for all three pages
PAGE_1_COORDINATES = {
    1:  [(798, 774),  (948, 774),  (1100, 774),  (1248, 774),  (1400, 774)],
    2:  [(798, 874),  (948, 874),  (1100, 874),  (1248, 874),  (1400, 874)],
    3:  [(798, 974),  (948, 974),  (1100, 974),  (1248, 974),  (1400, 974)],
    4:  [(798, 1041), (948, 1041), (1100, 1041), (1248, 1041), (1400, 1041)],
    5:  [(798, 1141), (948, 1141), (1100, 1141), (1248, 1141), (1400, 1141)],
    6:  [(798, 1241), (948, 1241), (1100, 1241), (1248, 1241), (1400, 1241)],
    7:  [(798, 1341), (948, 1341), (1100, 1341), (1248, 1341), (1400, 1341)],
    8:  [(798, 1441), (948, 1441), (1100, 1441), (1248, 1441), (1400, 1441)],
    9:  [(798, 1574), (948, 1574), (1100, 1574), (1248, 1574), (1400, 1574)],
    10: [(798, 1674), (948, 1674), (1100, 1674), (1248, 1674), (1400, 1674)],
    11: [(798, 1774), (948, 1774), (1100, 1774), (1248, 1774), (1400, 1774)],
}

PAGE_2_COORDINATES = {
    12: [(798, 466),  (948, 466),  (1098, 466),  (1248, 466),  (1400, 466)],
    13: [(798, 566),  (948, 566),  (1098, 566),  (1248, 566),  (1400, 566)],
    14: [(798, 666),  (948, 666),  (1098, 666),  (1248, 666),  (1400, 666)],
    15: [(798, 733),  (948, 733),  (1098, 733),  (1248, 733),  (1400, 733)],
    16: [(798, 800),  (948, 800),  (1098, 800),  (1248, 800),  (1400, 800)],
    17: [(798, 900),  (948, 900),  (1098, 900),  (1248, 900),  (1400, 900)],
    18: [(798, 1000), (948, 1000), (1098, 1000), (1248, 1000), (1400, 1000)],
    19: [(798, 1100), (948, 1100), (1098, 1100), (1248, 1100), (1400, 1100)],
    20: [(798, 1166), (948, 1166), (1098, 1166), (1248, 1166), (1400, 1166)],
    21: [(798, 1266), (948, 1266), (1098, 1266), (1248, 1266), (1400, 1266)],
    22: [(798, 1365), (948, 1365), (1098, 1365), (1248, 1365), (1400, 1365)],
    23: [(798, 1433), (948, 1433), (1098, 1433), (1248, 1433), (1400, 1433)],
    24: [(798, 1500), (948, 1500), (1098, 1500), (1248, 1500), (1400, 1500)],
    25: [(798, 1600), (948, 1600), (1098, 1600), (1248, 1600), (1400, 1600)],
    26: [(798, 1700), (948, 1700), (1098, 1700), (1248, 1700), (1400, 1700)],
}

PAGE_3_COORDINATES = {
    27: [(798, 466),  (948, 466),  (1100, 466),  (1248, 466),  (1400, 466)],
    28: [(798, 566),  (948, 566),  (1100, 566),  (1248, 566),  (1400, 566)],
    29: [(798, 666),  (948, 666),  (1100, 666),  (1248, 666),  (1400, 666)],
    30: [(798, 761),  (948, 761),  (1100, 761),  (1248, 761),  (1400, 761)],
    31: [(798, 861),  (948, 861),  (1100, 861),  (1248, 861),  (1400, 861)],
    32: [(798, 961),  (948, 961),  (1100, 961),  (1248, 961),  (1400, 961)],
    33: [(798, 1028), (948, 1028), (1100, 1028), (1248, 1028), (1400, 1028)],
}

# Combine all coordinates
ALL_COORDINATES = {**PAGE_1_COORDINATES, **PAGE_2_COORDINATES, **PAGE_3_COORDINATES}

# Page mapping for questions
QUESTION_TO_PAGE = {}
for q in range(1, 12):
    QUESTION_TO_PAGE[q] = 1
for q in range(12, 27):
    QUESTION_TO_PAGE[q] = 2
for q in range(27, 34):
    QUESTION_TO_PAGE[q] = 3

# Initialize session state for interactive image
if 'numbers' not in st.session_state:
    st.session_state.numbers = [0] * 33
if 'current_page' not in st.session_state:
    st.session_state.current_page = 1


def find_question_at_click(x, y, page_num):
    """
    Find which question number the user clicked on based on coordinates.
    Returns (question_num, option_value) or (None, None) if no match.
    """
    # Determine question range for this page
    if page_num == 1:
        question_range = range(1, 12)
    elif page_num == 2:
        question_range = range(12, 27)
    elif page_num == 3:
        question_range = range(27, 34)
    else:
        return None, None
    
    # Check each question's coordinates
    for question_num in question_range:
        if question_num in ALL_COORDINATES:
            coords = ALL_COORDINATES[question_num]
            # coords is a list of 5 tuples, one for each option (1-5)
            for option_idx, (coord_x, coord_y) in enumerate(coords):
                # Check if click is within 60 pixels of this coordinate
                distance = ((x - coord_x) ** 2 + (y - coord_y) ** 2) ** 0.5
                if distance < 60:
                    return question_num, option_idx + 1
    
    return None, None


def create_annotated_page_image(image_bytes, page_num, numbers):
    """
    Create an annotated version of a page with all markers for that page's questions.
    
    Args:
        image_bytes: BytesIO object containing the image data
        page_num: Page number (1, 2, or 3)
        numbers: List of all 33 user input numbers
    
    Returns:
        PIL Image object with all markers for that page
    """
    img = Image.open(image_bytes).convert("RGB")
    draw = ImageDraw.Draw(img, 'RGBA')
    
    marker_color = (255, 0, 0)  # Red
    marker_radius = 40
    
    # Determine question range for this page
    if page_num == 1:
        question_range = range(1, 12)
    elif page_num == 2:
        question_range = range(12, 27)
    elif page_num == 3:
        question_range = range(27, 34)
    else:
        return img
    
    # Draw all markers for this page
    for question_num in question_range:
        selected_option = numbers[question_num - 1]  # Get the user's answer (1-5)
        
        if question_num in ALL_COORDINATES and 1 <= selected_option <= 5:
            coords = ALL_COORDINATES[question_num]
            x, y = coords[selected_option - 1]
            
            # Draw circle with semi-transparent fill
            draw.ellipse(
                [(x - marker_radius, y - marker_radius), (x + marker_radius, y + marker_radius)],
                outline=marker_color,
                width=5,
                fill=marker_color + (100,)
            )
    
    return img


def process_excel_file(numbers, combined_name):
    """
    Process and fill the Excel template with the user's numbers.
    Returns BytesIO object with the filled workbook.
    """
    try:
        # Download template from GitHub
        github_url = "https://raw.githubusercontent.com/marianos-arch/adol-text-app/main/template.xlsx"
        response = requests.get(github_url)
        
        if response.status_code != 200:
            st.error(f"❌ Failed to download template. Status code: {response.status_code}")
            return None
        
        template_bytes = BytesIO(response.content)
        
        # Load workbook
        wb = load_workbook(template_bytes)
        wb.calculation.calcMode = 'auto' 
        ws = wb.active  # use the active sheet
        
        # Step 4: Map each question number to its specific cell
        # Format: question_number -> (row, col)
        question_cell_mapping = {
            1: (48, 4),   # D48
            2: (38, 4),   # D38
            3: (39, 4),   # D39
            4: (49, 4),   # D49
            5: (40, 4),   # D40
            6: (41, 4),   # D41
            7: (50, 4),   # D50
            8: (51, 4),   # D51
            9: (42, 4),   # D42
            10: (52, 4),  # D52
            11: (43, 4),  # D43
            12: (4, 4),   # D4
            13: (29, 4),  # D29
            14: (20, 4),  # D20
            15: (5, 4),   # D5
            16: (6, 4),   # D6
            17: (7, 4),   # D7
            18: (21, 4),  # D21
            19: (30, 4),  # D30
            20: (8, 4),   # D8
            21: (9, 4),   # D9
            22: (22, 4),  # D22
            23: (10, 4),  # D10
            24: (11, 4),  # D11
            25: (31, 4),  # D31
            26: (12, 4),  # D12
            27: (32, 4),  # D32
            28: (23, 4),  # D23
            29: (13, 4),  # D13
            30: (14, 4),  # D14
            31: (24, 4),  # D24
            32: (15, 4),  # D15
            33: (33, 4),  # D33
        }
        
        # Insert combined mentee name and date at C1
        ws.cell(row=1, column=3).value = combined_name
        
        # Insert each number into its corresponding cell
        for question_num, number in enumerate(numbers, start=1):
            if question_num in question_cell_mapping:
                row, col = question_cell_mapping[question_num]
                cell = ws.cell(row=row, column=col)
                cell.value = number

        #force Excel to recalculate formulas
        try:
            wb.calculate()
        except: 
                pass # some older excel versions may not support this
        
        # Step 5: Save to BytesIO for download
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
        
    except Exception as e:
        st.error(f"❌ Error processing template: {e}")
        st.info("Make sure the template.xlsx file exists in your GitHub repository.")
        return None


# INPUT METHOD 1: Type Numbers
if input_method == "Type Numbers":
    st.subheader("2. Enter Your Numbers (From Question 1 to Question 33)")
    user_input = st.text_input(
        "Enter numbers (e.g., '44442145' or '4 4 4 4 2 1 4 5'):",
        placeholder="44442145"
    )
    st.caption("Each number must be between 1 and 5. Strong Disagree (1) - Strongly Agree (5) ")

    # Step 3: Process and validate input
    if user_input:
        # Remove spaces and extract only digits
        cleaned_input = re.sub(r'\s+', '', user_input)
        numbers = [int(char) for char in cleaned_input if char.isdigit()]
        
        # Validate: Should have exactly 33 numbers (1-5 scale)
        if len(numbers) != 33:
            st.error(f"❌ Please enter exactly 33 numbers. You entered {len(numbers)}.")
        elif not all(1 <= num <= 5 for num in numbers):
            st.error("❌ All numbers must be between 1 and 5.")
        else:
            st.success(f"✅ Valid input! You entered: {numbers}")
            st.session_state.numbers = numbers
            
            output = process_excel_file(numbers, combined_name)
            
            if output:
                # Step 6: Provide download button
                # Generate filename from mentee name and date
                safe_name = re.sub(r'[^a-zA-Z0-9_-]', '', mentee_name.replace(' ', '_'))
                safe_date = re.sub(r'[^0-9_-]', '', date_input.replace('/', '_'))
                file_name = f"ADOL_Scoresheet_{safe_name}_{safe_date}.xlsx" if safe_name else "ADOL_Scoresheet_Filled.xlsx"

                st.download_button(
                        label="📥 Download Filled ADOL Sheet",
                        data=output.getvalue(),
                        file_name=file_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

                st.write(f"NOTE: Hit [Enable Editing] on Excel to reveal the scores") 
                
                # Display the numbers for verification with visual representation
                st.subheader("3. Your Entries - Visual Confirmation on Images:")
                st.write(f"Mentee Name and Today's Date: {combined_name}")
                
                # Display annotated images for each page
                # Try to load and display PNG images with markers
                try:
                    for page_num in [1, 2, 3]:
                        image_filename = f"adol_blank-{page_num}.png"
                        github_image_url = f"https://raw.githubusercontent.com/marianos-arch/adol-text-app/main/{image_filename}"
                        
                        try:
                            # Download image from GitHub
                            img_response = requests.get(github_image_url)
                            if img_response.status_code == 200:
                                # Create annotated version
                                annotated_img = create_annotated_page_image(
                                    BytesIO(img_response.content), 
                                    page_num, 
                                    numbers
                                )
                                
                                # Display the annotated image
                                if page_num == 1:
                                    st.markdown(f"**Page {page_num} - Questions 1-11:**")
                                elif page_num == 2:
                                    st.markdown(f"**Page {page_num} - Questions 12-26:**")
                                else:
                                    st.markdown(f"**Page {page_num} - Questions 27-33:**")
                                st.image(annotated_img, use_column_width=True)
                            else:
                                st.warning(f"⚠️ Could not load {image_filename} from GitHub.")
                        except Exception as e:
                            st.warning(f"⚠️ Error processing {image_filename}: {e}")
                
                except Exception as e:
                    st.info(f"Note: Image visualization requires PNG files in your repository: adol_blank-1.png, adol_blank-2.png, adol_blank-3.png")


# INPUT METHOD 2: Click on Image
else:
    st.subheader("2. Click on the ADOL Scoresheet to Enter Your Scores")
    st.write("Navigate through each page and click on your response for each question (1-5 scale)")
    
    # Try to load and display interactive PNG images
    try:
        # Display current page
        image_filename = f"adol_blank-{st.session_state.current_page}.png"
        github_image_url = f"https://raw.githubusercontent.com/marianos-arch/adol-text-app/main/{image_filename}"
        
        try:
            # Download image from GitHub
            img_response = requests.get(github_image_url)
            if img_response.status_code == 200:
                # Create annotated version
                annotated_img = create_annotated_page_image(
                    BytesIO(img_response.content), 
                    st.session_state.current_page, 
                    st.session_state.numbers
                )
                
                # Display the annotated image
                page_num = st.session_state.current_page
                if page_num == 1:
                    st.markdown(f"**Page {page_num} - Questions 1-11:**")
                elif page_num == 2:
                    st.markdown(f"**Page {page_num} - Questions 12-26:**")
                else:
                    st.markdown(f"**Page {page_num} - Questions 27-33:**")
                
                # Use streamlit-image-coordinates for clickable image
                if streamlit_image_coordinates is not None:
                    coords = streamlit_image_coordinates(annotated_img, key=f"coords_page_{page_num}")
                    
                    # Process click if coordinates were returned
                    if coords:
                        x = coords["x"]
                        y = coords["y"]
                        question_num, option_value = find_question_at_click(x, y, page_num)
                        
                        if question_num and option_value:
                            st.session_state.numbers[question_num - 1] = option_value
                            st.success(f"✅ Question {question_num} set to {option_value}")
                            st.rerun()
                        else:
                            st.info("Click on a response option to select it.")
                else:
                    st.image(annotated_img, use_column_width=True)
                    st.info("💡 Tip: Install streamlit-image-coordinates for clickable image support.")
                
                # Manual input for each question on current page
                if page_num == 1:
                    question_range = range(1, 12)
                elif page_num == 2:
                    question_range = range(12, 27)
                else:
                    question_range = range(27, 34)
                
                st.write("---")
                st.write("**Or enter scores manually for this page:**")
                
                # Create columns for input
                cols = st.columns(len(list(question_range)))
                for idx, question_num in enumerate(question_range):
                    with cols[idx]:
                        st.session_state.numbers[question_num - 1] = st.selectbox(
                            f"Q{question_num}",
                            options=[0, 1, 2, 3, 4, 5],
                            index=st.session_state.numbers[question_num - 1],
                            key=f"question_{question_num}"
                        )
                
                # Navigation buttons
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    if st.button("⬅️ Previous Page", disabled=(st.session_state.current_page == 1)):
                        st.session_state.current_page -= 1
                        st.rerun()
                
                with col2:
                    st.write(f"Page {st.session_state.current_page} of 3")
                
                with col3:
                    if st.button("Next Page ➡️", disabled=(st.session_state.current_page == 3)):
                        st.session_state.current_page += 1
                        st.rerun()
                
                # Show progress
                with col4:
                    filled_count = sum(1 for n in st.session_state.numbers if n > 0)
                    st.metric("Progress", f"{filled_count}/33")
                
                st.write("---")
                
                # Show all entries summary
                if any(n > 0 for n in st.session_state.numbers):
                    st.subheader("Your Current Entries:")
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        st.write("**Summary:**")
                        entries_str = ''.join(str(n) if n > 0 else '_' for n in st.session_state.numbers)
                        st.code(entries_str, language="text")
                    
                    with col2:
                        st.write("**Detailed View:**")
                        filled_answers = {i+1: n for i, n in enumerate(st.session_state.numbers) if n > 0}
                        st.json(filled_answers)
                    
                    # Check if all questions are answered
                    if all(n > 0 for n in st.session_state.numbers):
                        st.success("✅ All questions answered!")
                        
                        output = process_excel_file(st.session_state.numbers, combined_name)
                        
                        if output:
                            # Step 6: Provide download button
                            # Generate filename from mentee name and date
                            safe_name = re.sub(r'[^a-zA-Z0-9_-]', '', mentee_name.replace(' ', '_'))
                            safe_date = re.sub(r'[^0-9_-]', '', date_input.replace('/', '_'))
                            file_name = f"ADOL_Scoresheet_{safe_name}_{safe_date}.xlsx" if safe_name else "ADOL_Scoresheet_Filled.xlsx"

                            st.download_button(
                                    label="📥 Download Filled ADOL Sheet",
                                    data=output.getvalue(),
                                    file_name=file_name,
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )

                            st.write(f"NOTE: Hit [Enable Editing] on Excel to reveal the scores")
                    else:
                        st.info(f"📝 Answer all {sum(1 for n in st.session_state.numbers if n == 0)} remaining questions to enable download.")
                
            else:
                st.warning(f"⚠️ Could not load {image_filename} from GitHub.")
        except Exception as e:
            st.warning(f"⚠️ Error processing {image_filename}: {e}")
            st.write("Falling back to manual entry mode...")
            
            # Fallback: Manual entry
            st.write("Enter your scores for all 33 questions:")
            cols = st.columns(5)
            for q in range(1, 34):
                col_idx = (q - 1) % 5
                with cols[col_idx]:
                    st.session_state.numbers[q - 1] = st.selectbox(
                        f"Q{q}",
                        options=[0, 1, 2, 3, 4, 5],
                        index=st.session_state.numbers[q - 1],
                        key=f"question_{q}"
                    )
    
    except Exception as e:
        st.info(f"Note: Image visualization requires PNG files in your repository: adol_blank-1.png, adol_blank-2.png, adol_blank-3.png")
